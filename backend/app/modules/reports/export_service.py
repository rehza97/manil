"""
Export Service

Handles data exports in CSV, PDF, and Excel formats.
PDF can be generated from ReportLab (legacy) or from HTML template (HTML-to-PDF).
"""

import csv
import io
import os
from datetime import datetime, timezone
from typing import List, Dict, Any, Optional
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from weasyprint import HTML as WeasyHTML
    HTML_PDF_ENGINE = "weasyprint"
except ImportError:
    try:
        from xhtml2pdf import pisa
        HTML_PDF_ENGINE = "xhtml2pdf"
    except ImportError:
        HTML_PDF_ENGINE = None

from jinja2 import Environment, FileSystemLoader, select_autoescape

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfgen import canvas


class ExportService:
    """Service for exporting data in various formats."""

    def __init__(self, export_dir: str = "exports"):
        """
        Initialize export service.

        Args:
            export_dir: Directory to store exported files
        """
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(parents=True, exist_ok=True)
        templates_dir = Path(__file__).resolve().parent.parent.parent / "templates"
        self._jinja_env = Environment(
            loader=FileSystemLoader(str(templates_dir)),
            autoescape=select_autoescape(["html", "xml"]),
        ) if templates_dir.exists() else None

    def _html_to_pdf(self, html_content: str, output_path: str) -> None:
        """Convert HTML string to PDF using WeasyPrint or xhtml2pdf."""
        if HTML_PDF_ENGINE == "weasyprint":
            doc = WeasyHTML(string=html_content)
            doc.write_pdf(output_path)
        elif HTML_PDF_ENGINE == "xhtml2pdf":
            with open(output_path, "wb") as f:
                status = pisa.CreatePDF(html_content, dest=f)
                if status.err:
                    raise RuntimeError("HTML-to-PDF conversion failed")
        else:
            raise RuntimeError(
                "No HTML-to-PDF engine. Install weasyprint or xhtml2pdf."
            )

    def export_to_csv(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        headers: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """
        Export data to CSV format

        Args:
            data: List of dictionaries containing data
            filename: Output filename (without extension)
            headers: Optional list of column headers

        Returns:
            Dictionary with file information
        """
        if headers is None and not data:
            raise ValueError("No data to export and no headers provided")

        # Generate filename
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        file_name = f"{filename}_{timestamp}.csv"
        file_path = self.export_dir / file_name

        # Determine headers
        if headers is None:
            headers = list(data[0].keys())

        # Write CSV
        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=headers)
            writer.writeheader()
            writer.writerows(data)

        return {
            "file_name": file_name,
            "file_path": str(file_path),
            "file_size": os.path.getsize(file_path),
            "format": "csv",
            "generated_at": datetime.now(timezone.utc)
        }

    def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        sheet_name: str = "Report",
        headers: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """
        Export data to Excel format

        Args:
            data: List of dictionaries containing data
            filename: Output filename (without extension)
            sheet_name: Name of the Excel sheet
            headers: Optional list of column headers

        Returns:
            Dictionary with file information
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        if headers is None and not data:
            raise ValueError("No data to export and no headers provided")

        # Generate filename
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        file_name = f"{filename}_{timestamp}.xlsx"
        file_path = self.export_dir / file_name

        # Determine headers
        if headers is None:
            headers = list(data[0].keys())

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Style for header row
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write data
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, header in enumerate(headers, start=1):
                value = row_data.get(header, "")
                # Convert datetime objects to strings
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M:%S")
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(file_path)

        return {
            "file_name": file_name,
            "file_path": str(file_path),
            "file_size": os.path.getsize(file_path),
            "format": "excel",
            "generated_at": datetime.now(timezone.utc)
        }

    def export_to_pdf(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        title: str,
        headers: Optional[List[str]] = None,
        include_summary: bool = True
    ) -> Dict[str, Any]:
        """
        Export data to PDF format

        Args:
            data: List of dictionaries containing data
            filename: Output filename (without extension)
            title: Report title
            headers: Optional list of column headers
            include_summary: Whether to include a summary section

        Returns:
            Dictionary with file information
        """
        if headers is None and not data:
            raise ValueError("No data to export and no headers provided")

        # Generate filename
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        file_name = f"{filename}_{timestamp}.pdf"
        file_path = self.export_dir / file_name

        # Create PDF
        doc = SimpleDocTemplate(str(file_path), pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1F4788'),
            spaceAfter=30,
            alignment=1  # Center
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 12))

        # Metadata
        meta_style = styles['Normal']
        generated_text = f"Generated on: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}"
        elements.append(Paragraph(generated_text, meta_style))
        elements.append(Spacer(1, 12))

        # Summary section
        if include_summary:
            summary_style = ParagraphStyle(
                'Summary',
                parent=styles['Heading2'],
                fontSize=14,
                textColor=colors.HexColor('#1F4788'),
                spaceAfter=12
            )
            elements.append(Paragraph("Summary", summary_style))
            elements.append(Paragraph(f"Total Records: {len(data)}", meta_style))
            elements.append(Spacer(1, 20))

        # Determine headers
        if headers is None:
            headers = list(data[0].keys())

        # Prepare table data
        table_data = [headers]

        for row in data:
            table_row = []
            for header in headers:
                value = row.get(header, "")
                # Convert datetime objects to strings
                if isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M")
                # Truncate long strings
                value_str = str(value)
                if len(value_str) > 50:
                    value_str = value_str[:47] + "..."
                table_row.append(value_str)
            table_data.append(table_row)

        # Create table
        table = Table(table_data)

        # Table style
        table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Data rows
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
        ]))

        elements.append(table)

        # Build PDF
        doc.build(elements)

        return {
            "file_name": file_name,
            "file_path": str(file_path),
            "file_size": os.path.getsize(file_path),
            "format": "pdf",
            "generated_at": datetime.now(timezone.utc)
        }

    def export_to_pdf_html(
        self,
        data: List[Dict[str, Any]],
        filename: str,
        title: str,
        headers: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        """
        Export data to PDF by rendering HTML template then converting to PDF.
        Falls back to ReportLab (export_to_pdf) if HTML path fails.
        """
        if headers is None and not data:
            raise ValueError("No data to export and no headers provided")
        if headers is None:
            headers = list(data[0].keys())

        if HTML_PDF_ENGINE is not None and self._jinja_env is not None:
            try:
                rows = []
                for row in data:
                    cells = []
                    for h in headers:
                        value = row.get(h, "")
                        if isinstance(value, datetime):
                            value = value.strftime("%Y-%m-%d %H:%M")
                        value_str = str(value)
                        if len(value_str) > 50:
                            value_str = value_str[:47] + "..."
                        cells.append(value_str)
                    rows.append(cells)

                generated_at = datetime.now(timezone.utc).strftime(
                    "%Y-%m-%d %H:%M:%S UTC"
                )
                timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
                file_name = f"{filename}_{timestamp}.pdf"
                file_path = self.export_dir / file_name

                template = self._jinja_env.get_template("reports/report_pdf.html")
                html_content = template.render(
                    title=title,
                    generated_at=generated_at,
                    total_records=len(data),
                    headers=headers,
                    rows=rows,
                )
                self._html_to_pdf(html_content, str(file_path))

                return {
                    "file_name": file_name,
                    "file_path": str(file_path),
                    "file_size": os.path.getsize(file_path),
                    "format": "pdf",
                    "generated_at": datetime.now(timezone.utc),
                }
            except Exception:
                pass

        return self.export_to_pdf(data, filename, title, headers)

    def export_tickets_report(
        self,
        tickets: List[Dict[str, Any]],
        format: str = "csv",
        report_name: str = "tickets_report"
    ) -> Dict[str, Any]:
        """
        Export tickets report in specified format

        Args:
            tickets: List of ticket dictionaries
            format: Export format (csv, excel, pdf)
            report_name: Name for the export file

        Returns:
            Dictionary with file information
        """
        headers = [
            "id", "subject", "status", "priority", "created_at",
            "customer_id", "assigned_to_id", "first_response_at", "resolved_at"
        ]

        if format == "csv":
            return self.export_to_csv(tickets, report_name, headers)
        elif format == "excel":
            return self.export_to_excel(tickets, report_name, "Tickets", headers)
        elif format == "pdf":
            return self.export_to_pdf(tickets, report_name, "Tickets Report", headers)
        else:
            raise ValueError(f"Unsupported format: {format}")

    def export_customers_report(
        self,
        customers: List[Dict[str, Any]],
        format: str = "csv",
        report_name: str = "customers_report"
    ) -> Dict[str, Any]:
        """
        Export customers report in specified format

        Args:
            customers: List of customer dictionaries
            format: Export format (csv, excel, pdf)
            report_name: Name for the export file

        Returns:
            Dictionary with file information
        """
        headers = [
            "id", "full_name", "email", "phone", "customer_type",
            "status", "created_at"
        ]

        if format == "csv":
            return self.export_to_csv(customers, report_name, headers)
        elif format == "excel":
            return self.export_to_excel(customers, report_name, "Customers", headers)
        elif format == "pdf":
            return self.export_to_pdf(customers, report_name, "Customers Report", headers)
        else:
            raise ValueError(f"Unsupported format: {format}")

    def export_orders_report(
        self,
        orders: List[Dict[str, Any]],
        format: str = "csv",
        report_name: str = "orders_report"
    ) -> Dict[str, Any]:
        """
        Export orders report in specified format

        Args:
            orders: List of order dictionaries
            format: Export format (csv, excel, pdf)
            report_name: Name for the export file

        Returns:
            Dictionary with file information
        """
        headers = [
            "id", "order_number", "customer_id", "status",
            "subtotal_amount", "tax_amount", "total_amount", "created_at"
        ]

        if format == "csv":
            return self.export_to_csv(orders, report_name, headers)
        elif format == "excel":
            return self.export_to_excel(orders, report_name, "Orders", headers)
        elif format == "pdf":
            return self.export_to_pdf(orders, report_name, "Orders Report", headers)
        else:
            raise ValueError(f"Unsupported format: {format}")

    def export_vps_report(
        self,
        vps_rows: List[Dict[str, Any]],
        format: str = "csv",
        report_name: str = "vps_report"
    ) -> Dict[str, Any]:
        """
        Export VPS (hosting) subscriptions report in specified format.

        Args:
            vps_rows: List of VPS subscription dicts (id, subscription_number, etc.)
            format: Export format (csv, excel, pdf)
            report_name: Name for the export file

        Returns:
            Dictionary with file information
        """
        headers = [
            "id", "subscription_number", "customer_id", "customer_name", "customer_email",
            "plan_name", "plan_slug", "plan_cpu_cores", "plan_ram_gb", "plan_storage_gb",
            "status", "monthly_price", "created_at"
        ]
        if format == "csv":
            return self.export_to_csv(vps_rows, report_name, headers)
        elif format == "excel":
            return self.export_to_excel(vps_rows, report_name, "VPS Subscriptions", headers)
        elif format == "pdf":
            return self.export_to_pdf(vps_rows, report_name, "VPS Report", headers)
        else:
            raise ValueError(f"Unsupported format: {format}")
