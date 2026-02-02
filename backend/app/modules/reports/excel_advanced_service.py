"""
Advanced Excel Service

Multi-sheet Excel reports: Summary, Details, Raw Data; header styling, number formats.
"""

import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Any, List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class AdvancedExcelService:
    """Multi-sheet Excel reports with formatting."""

    HEADER_FILL = "1F4788"
    HEADER_FONT_COLOR = "FFFFFF"

    def __init__(self, export_dir: Optional[str] = None):
        self.export_dir = Path(export_dir or "exports")
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def create_multi_sheet_report(
        self,
        report_data: Dict[str, Any],
        filename: str,
    ) -> str:
        """
        Create Excel with Summary, Details, Raw Data sheets.
        report_data: summary (dict), details (list of dicts), raw (list of dicts).
        Returns absolute file path.
        """
        if not EXCEL_AVAILABLE:
            raise ImportError(
                "openpyxl is required. Install with: pip install openpyxl")
        ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        file_name = f"{filename}_{ts}.xlsx"
        file_path = self.export_dir / file_name
        wb = Workbook()
        # Summary
        ws_sum = wb.active
        ws_sum.title = "Summary"
        self._write_sheet_from_dict(ws_sum, report_data.get("summary") or {})
        # Details
        ws_det = wb.create_sheet("Details")
        self._write_sheet_from_rows(ws_det, report_data.get("details") or [])
        # Raw Data
        ws_raw = wb.create_sheet("Raw Data")
        self._write_sheet_from_rows(ws_raw, report_data.get(
            "raw") or report_data.get("details") or [])
        self._apply_header_style(wb)
        wb.save(str(file_path))
        return str(file_path.resolve())

    def _write_sheet_from_dict(self, ws, data: Dict[str, Any]) -> None:
        for i, (k, v) in enumerate(data.items(), start=1):
            ws.cell(row=i, column=1, value=str(k))
            ws.cell(row=i, column=2, value=v)

    def _write_sheet_from_rows(self, ws, rows: List[Dict[str, Any]]) -> None:
        if not rows:
            return
        headers = list(rows[0].keys())
        for col, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=h)
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, h in enumerate(headers, start=1):
                val = row.get(h)
                if isinstance(val, datetime):
                    val = val.strftime("%Y-%m-%d %H:%M:%S")
                ws.cell(row=row_idx, column=col_idx, value=val)

    def _apply_header_style(self, wb: "Workbook") -> None:
        fill = PatternFill(start_color=self.HEADER_FILL,
                           end_color=self.HEADER_FILL, fill_type="solid")
        font = Font(bold=True, color=self.HEADER_FONT_COLOR, size=12)
        for ws in wb.worksheets:
            for cell in ws[1]:
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(
                    horizontal="center", vertical="center")
            for column in ws.columns:
                max_len = max(len(str(c.value or "")) for c in column)
                ws.column_dimensions[column[0].column_letter].width = min(
                    max_len + 2, 50)
